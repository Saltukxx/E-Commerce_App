"""
fetch_images.py  — runs on Windows Python
1. Reads duzgun_urun_listesi_20260509.xlsx
2. Translates Turkish product names -> German (Google Translate)
3. Searches DuckDuckGo Images for each product
4. Downloads best image into product-images/<lagercode>.jpg
5. Saves urunler_mit_bildern_DE.xlsx with German names + image paths
"""

import os, time, random, hashlib, re, sys

import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from deep_translator import GoogleTranslator
from ddgs import DDGS

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE     = r"C:\Users\satog\OneDrive\Desktop\ecommercefork\export"
EXCEL_IN  = os.path.join(BASE, "duzgun_urun_listesi_20260509.xlsx")
EXCEL_OUT = os.path.join(BASE, "urunler_mit_bildern_DE.xlsx")
IMG_DIR   = os.path.join(BASE, "product-images")
LOG_FILE  = os.path.join(BASE, "fetch_images_progress.txt")

os.makedirs(IMG_DIR, exist_ok=True)

def log(msg):
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        print(msg.encode('ascii', errors='replace').decode(), flush=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

# ─── Helpers ──────────────────────────────────────────────────────────────────
translator = GoogleTranslator(source='tr', target='de')

def translate(text):
    try:
        return translator.translate(text) or text
    except Exception as e:
        log(f"    [translate error] {e}")
        return text

def safe_filename(lagercode, title):
    code = re.sub(r'[^\w\-]', '_', lagercode) if lagercode else ''
    if not code:
        code = hashlib.md5(title.encode()).hexdigest()[:8]
    return code

def search_image(query):
    try:
        ddgs = DDGS()
        results = list(ddgs.images(query, max_results=4))
        for r in results:
            url = r.get('image', '')
            if url and any(ext in url.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp']):
                return url
        return results[0]['image'] if results else None
    except Exception as e:
        log(f"    [search error] {e}")
        return None

def download_image(url, filepath):
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
        }
        resp = requests.get(url, headers=headers, timeout=12, stream=True)
        ct = resp.headers.get('Content-Type', '')
        if resp.status_code == 200 and 'image' in ct:
            with open(filepath, 'wb') as f:
                for chunk in resp.iter_content(8192):
                    f.write(chunk)
            if os.path.getsize(filepath) < 1000:
                os.remove(filepath)
                return False
            return True
    except Exception as e:
        log(f"    [download error] {e}")
    return False

# ─── Load products ────────────────────────────────────────────────────────────
wb = load_workbook(EXCEL_IN)
ws = wb.active

products = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row[0]:
        continue
    products.append({
        'row_idx':   i,
        'title_tr':  str(row[0]).strip(),
        'brand':     str(row[1] or '').strip(),
        'category':  str(row[2] or '').strip(),
        'stock':     str(row[3] or '').strip(),
        'einkauf':   row[4],
        'liste':     row[5],
        'verkauf':   row[6],
        'minimum':   row[7],
        'lagercode': str(row[8] or '').strip(),
    })

total = len(products)
log(f"Loaded {total} products.\n" + "=" * 70)

results = []
for idx, p in enumerate(products, 1):
    title_tr  = p['title_tr']
    lagercode = p['lagercode']
    category  = p['category']
    filename  = safe_filename(lagercode, title_tr)
    img_path  = os.path.join(IMG_DIR, filename + '.jpg')
    rel_path  = f"product-images/{filename}.jpg"

    log(f"[{idx:>3}/{total}] {title_tr[:60]}")

    # Translate
    title_de = translate(title_tr)
    log(f"        DE: {title_de[:65]}")

    # Already downloaded?
    if os.path.exists(img_path):
        log(f"        >> cached")
        results.append({**p, 'title_de': title_de, 'img_path': rel_path, 'img_status': 'cached'})
        continue

    # Search
    img_url = search_image(f"{title_de} {category}")
    if not img_url:
        short = ' '.join(title_de.split()[:5])
        img_url = search_image(short)

    # Download
    status = 'not found'
    if img_url:
        ok = download_image(img_url, img_path)
        status = 'ok' if ok else 'failed'
        log(f"        >> {status}: {img_url[:80]}")
    else:
        log(f"        >> no image found")

    results.append({
        **p,
        'title_de':   title_de,
        'img_path':   rel_path if status == 'ok' else '',
        'img_status': status,
    })

    time.sleep(random.uniform(0.9, 1.6))

# ─── Write output Excel ───────────────────────────────────────────────────────
log("\nWriting Excel output...")
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Produkte DE'

hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(color="FFFFFF", bold=True)

headers = ['Lagercode','Artikel (TR)','Artikel (DE)','Marke','Kategorie',
           'Bestand','Einkauf','Liste','Verkauf','Minimum','Bild-Datei','Status']
ws_out.append(headers)
for cell in ws_out[1]:
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

col_widths = [12,55,55,14,22,10,11,11,11,10,38,12]
for i, w in enumerate(col_widths, 1):
    ws_out.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ok_fill   = PatternFill("solid", fgColor="E2EFDA")
miss_fill = PatternFill("solid", fgColor="FCE4D6")

for r in results:
    ws_out.append([
        r['lagercode'], r['title_tr'], r['title_de'], r['brand'], r['category'],
        r['stock'], r['einkauf'], r['liste'], r['verkauf'], r['minimum'],
        r['img_path'], r['img_status'],
    ])
    fill = ok_fill if r['img_status'] in ('ok','cached') else miss_fill
    for cell in ws_out[ws_out.max_row]:
        cell.fill = fill

wb_out.save(EXCEL_OUT)

ok_n   = sum(1 for r in results if r['img_status'] in ('ok','cached'))
fail_n = len(results) - ok_n
log(f"\n{'='*70}")
log(f"  Images downloaded : {ok_n}")
log(f"  Failed/not found  : {fail_n}")
log(f"  Excel saved       : {EXCEL_OUT}")
log(f"  Images folder     : {IMG_DIR}")
log("DONE.")
