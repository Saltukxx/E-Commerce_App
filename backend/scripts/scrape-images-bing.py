"""
scrape-images-bing.py

Fetches one product image per item from Bing Image Search API and saves
the results to image-results.json, ready for update-images.ts to push to DB.

HOW TO RUN:
  cd backend
  pip install requests openpyxl          # one-time
  python scripts/scrape-images-bing.py   # set BING_API_KEY below first

GET A FREE KEY (1,000 calls/month free, then $3 / 1,000):
  https://portal.azure.com → Create resource → search "Bing Search v7"
  After creating, copy Key 1 from the resource's "Keys and Endpoint" page.

RESUMABLE: safe to re-run — already-scraped codes are skipped.
"""

import os
import json
import time
import requests
import openpyxl
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

BING_API_KEY = os.environ.get("BING_API_KEY", "PASTE_YOUR_KEY_HERE")

EXCEL_PATH   = Path(__file__).parent.parent.parent / "Refrigeration Catalogue 2025 - Product Data (1).xlsx"
RESULTS_PATH = Path(__file__).parent / "image-results.json"
PROGRESS_PATH = Path(__file__).parent / "scrape-progress.json"

BING_ENDPOINT = "https://api.bing.microsoft.com/v7.0/images/search"
DELAY_SECONDS = 0.4          # ~2.5 req/sec — well within rate limits
SAVE_EVERY    = 25            # checkpoint frequency

# ── Helpers ───────────────────────────────────────────────────────────────────

def build_query(model: str, category: str, repa_it: str) -> str:
    """Build a search query that finds the right product image."""
    # Clean up the category to extract the brand/type
    # e.g. "Cubigel compressors" → "Cubigel"
    # e.g. "Embraco compressors" → "Embraco"
    # e.g. "Electronic controllers" → just use model
    cat_clean = category.replace(" compressors", "").replace(" COMPRESSORS", "").strip()

    if model and model != "nan":
        return f"{cat_clean} {model} refrigeration"
    else:
        return f"{repa_it} {cat_clean} refrigeration spare part"


def fetch_image(query: str, api_key: str) -> dict | None:
    """Call Bing Image Search and return the best result."""
    headers = {"Ocp-Apim-Subscription-Key": api_key}
    params  = {
        "q":           query,
        "count":       5,
        "imageType":   "Photo",
        "safeSearch":  "Strict",
        # Prefer larger images
        "minWidth":    200,
        "minHeight":   200,
    }
    try:
        r = requests.get(BING_ENDPOINT, headers=headers, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        values = data.get("value", [])
        if not values:
            return None
        # Return the first result (Bing already ranks by relevance)
        best = values[0]
        return {
            "imageUrl":    best.get("contentUrl"),
            "thumbnailUrl": best.get("thumbnailUrl"),
            "name":        best.get("name"),
            "hostDomain":  best.get("hostPageDomainFriendlyName"),
        }
    except Exception as e:
        return {"error": str(e)}


def load_excel() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["All Products"]
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        repa_it  = str(row_dict.get("REPA IT Code", "") or "").strip()
        if not repa_it or repa_it in ("None", "nan", ""):
            continue
        products.append({
            "repaItCode": repa_it,
            "model":      str(row_dict.get("Model", "") or "").strip(),
            "category":   str(row_dict.get("Category", "") or "").strip(),
        })
    wb.close()
    return products


def load_progress() -> set:
    if not PROGRESS_PATH.exists():
        return set()
    return set(json.loads(PROGRESS_PATH.read_text()))


def save_progress(done: set):
    PROGRESS_PATH.write_text(json.dumps(list(done), indent=2))


def load_results() -> dict:
    if not RESULTS_PATH.exists():
        return {}
    return {r["repaItCode"]: r for r in json.loads(RESULTS_PATH.read_text())}


def save_results(results: dict):
    RESULTS_PATH.write_text(json.dumps(list(results.values()), indent=2))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if BING_API_KEY == "PASTE_YOUR_KEY_HERE":
        print("ERROR: Set your Bing API key in the script or via:")
        print("  export BING_API_KEY=your_key_here")
        print("  python scripts/scrape-images-bing.py")
        return

    print("Loading Excel...")
    products = load_excel()
    print(f"Found {len(products)} products.")

    done    = load_progress()
    results = load_results()
    remaining = [p for p in products if p["repaItCode"] not in done]

    print(f"Already done: {len(done)}. Remaining: {len(remaining)}\n")

    if not remaining:
        print("All done! Check scripts/image-results.json")
        return

    last_save = time.time()

    for i, product in enumerate(remaining):
        repa_it  = product["repaItCode"]
        model    = product["model"]
        category = product["category"]
        total    = len(done) + i + 1
        query    = build_query(model, category, repa_it)

        print(f"[{total}/{len(products)}] {repa_it} ({model}) ← \"{query}\"")

        result = fetch_image(query, BING_API_KEY)

        if result and "error" not in result and result.get("imageUrl"):
            image_url = result["imageUrl"]
            print(f"  ✓ {image_url[:80]}...")
            results[repa_it] = {
                "repaItCode": repa_it,
                "model":      model,
                "imageUrl":   image_url,
                "query":      query,
                "status":     "found",
            }
        elif result and "error" in result:
            print(f"  ⚠ API error: {result['error']}")
            results[repa_it] = {
                "repaItCode": repa_it,
                "model":      model,
                "imageUrl":   None,
                "query":      query,
                "status":     "error",
                "error":      result["error"],
            }
        else:
            print(f"  ✗ No image found")
            results[repa_it] = {
                "repaItCode": repa_it,
                "model":      model,
                "imageUrl":   None,
                "query":      query,
                "status":     "not_found",
            }

        done.add(repa_it)

        # Checkpoint
        if (i + 1) % SAVE_EVERY == 0 or time.time() - last_save > 60:
            save_progress(done)
            save_results(results)
            last_save = time.time()
            print(f"  → Saved checkpoint ({len(done)} done)")

        time.sleep(DELAY_SECONDS)

    # Final save
    save_progress(done)
    save_results(results)

    found     = sum(1 for r in results.values() if r["status"] == "found")
    not_found = sum(1 for r in results.values() if r["status"] == "not_found")
    errors    = sum(1 for r in results.values() if r["status"] == "error")

    print("\n════════════════════════════════")
    print(f"  Total scraped : {len(results)}")
    print(f"  Images found  : {found}")
    print(f"  Not found     : {not_found}")
    print(f"  Errors        : {errors}")
    print("════════════════════════════════")
    print("\nNext step:  cd backend && npx ts-node --transpile-only scripts/update-images.ts")


if __name__ == "__main__":
    main()
