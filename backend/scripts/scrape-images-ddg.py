"""
scrape-images-ddg.py

Fetches one product image per item using DuckDuckGo image search.
No API key. No account. Completely free.

HOW TO RUN:
  cd backend
  pip install ddgs openpyxl
  python scripts/scrape-images-ddg.py

RESUMABLE: safe to re-run — already-scraped codes are skipped.
After this finishes, run:
  npx ts-node --transpile-only scripts/update-images.ts
"""

import json
import time
import random
import openpyxl
from pathlib import Path
from ddgs import DDGS

# ── Config ────────────────────────────────────────────────────────────────────

EXCEL_PATH    = Path(__file__).parent.parent.parent / "Refrigeration Catalogue 2025 - Product Data (1).xlsx"
RESULTS_PATH  = Path(__file__).parent / "image-results.json"
PROGRESS_PATH = Path(__file__).parent / "scrape-progress.json"

SAVE_EVERY    = 25       # checkpoint every N products
BASE_DELAY    = 1.2      # seconds between requests
JITTER        = 0.6      # random extra delay to avoid rate limits

# ── Helpers ───────────────────────────────────────────────────────────────────

def build_query(model: str, category: str, repa_it: str) -> str:
    """Build a focused search query from product fields."""
    # Strip generic suffixes from category to extract the useful brand/type name
    # e.g. "Cubigel compressors" → "Cubigel"
    # e.g. "Filters COPPER FILTERS" → "copper filter refrigeration"
    # e.g. "Electronic controllers" → just use model
    clean = (
        category
        .replace(" compressors", "").replace(" COMPRESSORS", "")
        .replace(" COMPRESSOR", "")
        .strip()
    )

    if model and model not in ("nan", "None", ""):
        return f"{clean} {model} refrigeration"
    return f"{repa_it} {clean} refrigeration spare part"


def fetch_image_ddg(query: str) -> str | None:
    """Search DuckDuckGo images and return the best image URL."""
    try:
        with DDGS() as ddgs:
            results = list(ddgs.images(query, max_results=5))
        if not results:
            return None
        # Prefer results from known supplier / manufacturer domains
        priority_domains = [
            "embraco", "secop", "tecumseh", "danfoss", "cubigel",
            "refrigeration", "compressor", "hvac", "spare", "part",
        ]
        for r in results:
            url = r.get("image", "")
            if any(d in url.lower() for d in priority_domains):
                return url
        # Fall back to first result
        return results[0].get("image")
    except Exception as e:
        return f"ERROR:{e}"


def load_excel() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["All Products"]
    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        code = str(d.get("REPA IT Code", "") or "").strip()
        if not code or code in ("None", "nan", ""):
            continue
        products.append({
            "repaItCode": code,
            "model":      str(d.get("Model", "") or "").strip(),
            "category":   str(d.get("Category", "") or "").strip(),
        })
    wb.close()
    return products


def load_progress() -> set:
    if not PROGRESS_PATH.exists():
        return set()
    return set(json.loads(PROGRESS_PATH.read_text()))


def save_progress(done: set):
    PROGRESS_PATH.write_text(json.dumps(list(done), indent=2))


def load_results() -> dict:
    if not RESULTS_PATH.exists():
        return {}
    return {r["repaItCode"]: r for r in json.loads(RESULTS_PATH.read_text())}


def save_results(results: dict):
    RESULTS_PATH.write_text(json.dumps(list(results.values()), indent=2))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading Excel...")
    products = load_excel()
    print(f"Found {len(products)} products.\n")

    done      = load_progress()
    results   = load_results()
    remaining = [p for p in products if p["repaItCode"] not in done]

    print(f"Already done : {len(done)}")
    print(f"Remaining    : {len(remaining)}\n")

    if not remaining:
        print("All products already scraped.")
        print(f"Results → {RESULTS_PATH}")
        return

    last_save = time.time()

    for i, product in enumerate(remaining):
        code     = product["repaItCode"]
        model    = product["model"]
        category = product["category"]
        query    = build_query(model, category, code)
        n        = len(done) + i + 1

        print(f"[{n}/{len(products)}] {code} ({model})")
        print(f"  query: \"{query}\"")

        url = fetch_image_ddg(query)

        if url and url.startswith("ERROR:"):
            print(f"  ⚠  {url}")
            results[code] = {
                "repaItCode": code, "model": model,
                "imageUrl": None, "query": query,
                "status": "error", "error": url[6:],
            }
        elif url:
            print(f"  ✓  {url[:90]}")
            results[code] = {
                "repaItCode": code, "model": model,
                "imageUrl": url, "query": query,
                "status": "found",
            }
        else:
            print(f"  ✗  no image found")
            results[code] = {
                "repaItCode": code, "model": model,
                "imageUrl": None, "query": query,
                "status": "not_found",
            }

        done.add(code)

        if (i + 1) % SAVE_EVERY == 0 or time.time() - last_save > 60:
            save_progress(done)
            save_results(results)
            last_save = time.time()
            print(f"  → checkpoint saved ({len(done)} done)\n")

        # Polite delay with jitter to avoid rate-limit bans
        time.sleep(BASE_DELAY + random.uniform(0, JITTER))

    save_progress(done)
    save_results(results)

    found     = sum(1 for r in results.values() if r["status"] == "found")
    not_found = sum(1 for r in results.values() if r["status"] == "not_found")
    errors    = sum(1 for r in results.values() if r["status"] == "error")

    print("\n════════════════════════════════════════")
    print(f"  Total   : {len(results)}")
    print(f"  Found   : {found}  ✓")
    print(f"  Missing : {not_found}  ✗")
    print(f"  Errors  : {errors}  ⚠")
    print("════════════════════════════════════════")
    print(f"\nResults saved → {RESULTS_PATH}")
    print("Next step:")
    print("  cd backend && npx ts-node --transpile-only scripts/update-images.ts")


if __name__ == "__main__":
    main()
